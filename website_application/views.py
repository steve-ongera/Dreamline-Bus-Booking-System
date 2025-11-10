from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Count, Prefetch
from django.utils import timezone
from django.core.cache import cache
from datetime import datetime, timedelta
from decimal import Decimal
import json

from .models import (
    Location, Trip, Seat, Booking, SeatBooking, 
    Payment, BoardingPoint, RouteStop
)


def home_view(request):
    """Home page with search form"""
    context = {
        'page_title': 'Book Your Bus Ticket'
    }
    return render(request, 'home.html', context)


def search_results_view(request):
    """Search results page showing available trips"""
    origin_id = request.GET.get('origin')
    destination_id = request.GET.get('destination')
    travel_date = request.GET.get('date')
    
    if not all([origin_id, destination_id, travel_date]):
        return render(request, 'search_results.html', {
            'error': 'Please provide origin, destination, and travel date'
        })
    
    try:
        origin = Location.objects.get(id=origin_id)
        destination = Location.objects.get(id=destination_id)
        date_obj = datetime.strptime(travel_date, '%Y-%m-%d').date()
    except (Location.DoesNotExist, ValueError):
        return render(request, 'search_results.html', {
            'error': 'Invalid search parameters'
        })
    
    context = {
        'origin': origin,
        'destination': destination,
        'travel_date': date_obj,
        'travel_date_str': travel_date,
        'page_title': f'{origin.name} to {destination.name}'
    }
    
    return render(request, 'search_results.html', context)


# ============== AJAX API VIEWS ==============

@require_http_methods(["GET"])
def api_autocomplete_locations(request):
    """Autocomplete API for location search"""
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    locations = Location.objects.filter(
        Q(name__icontains=query) | Q(county__icontains=query),
        is_active=True
    )[:10]
    
    results = [
        {
            'id': loc.id,
            'name': loc.name,
            'county': loc.county,
            'display': f"{loc.name}, {loc.county}"
        }
        for loc in locations
    ]
    
    return JsonResponse({'results': results})


@require_http_methods(["GET"])
def api_search_trips(request):
    """API to search for available trips"""
    origin_id = request.GET.get('origin')
    destination_id = request.GET.get('destination')
    travel_date = request.GET.get('date')
    
    if not all([origin_id, destination_id, travel_date]):
        return JsonResponse({'error': 'Missing parameters'}, status=400)
    
    try:
        date_obj = datetime.strptime(travel_date, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date format'}, status=400)
    
    # Find direct routes and routes that pass through
    trips = Trip.objects.filter(
        Q(route__origin_id=origin_id, route__destination_id=destination_id) |
        (
            Q(route__stops__boarding_point__location_id=origin_id) &
            Q(route__stops__boarding_point__location_id=destination_id)
        ),
        departure_date=date_obj,
        is_active=True,
        status='scheduled'
    ).select_related(
        'bus__operator',
        'bus__seat_layout',
        'route__origin',
        'route__destination'
    ).prefetch_related(
        'bus__amenities',
        'seats'
    ).distinct().order_by('departure_time')

    
    results = []
    for trip in trips:
        # Count available seats by class
        seats = trip.seats.filter(is_available=True)
        vip_count = seats.filter(seat_class='vip').count()
        business_count = seats.filter(seat_class='business').count()
        normal_count = seats.filter(seat_class='normal').count()
        
        # Get amenities
        amenities = [
            {'name': amenity.name, 'icon': amenity.icon}
            for amenity in trip.bus.amenities.all()
        ]
        
        results.append({
            'id': trip.id,
            'bus_name': trip.bus.bus_name,
            'operator': trip.bus.operator.name,
            'route': f"{trip.route.origin.name} → {trip.route.destination.name}",
            'departure_time': trip.departure_time.strftime('%H:%M'),
            'arrival_time': trip.arrival_time.strftime('%H:%M'),
            'rating': float(trip.bus.rating),
            'total_ratings': trip.bus.total_ratings,
            'amenities': amenities,
            'prices': {
                'vip': float(trip.base_fare_vip) if vip_count > 0 else None,
                'business': float(trip.base_fare_business) if business_count > 0 else None,
                'normal': float(trip.base_fare_normal) if normal_count > 0 else None,
            },
            'available_seats': {
                'vip': vip_count,
                'business': business_count,
                'normal': normal_count,
                'total': vip_count + business_count + normal_count
            }
        })
    
    return JsonResponse({'trips': results})


@require_http_methods(["GET"])
def api_get_seats(request, trip_id):
    """API to get seat layout and availability for a trip"""
    trip = get_object_or_404(Trip, id=trip_id)
    
    # Get all seats with their status
    seats = trip.seats.all().order_by('row_number', 'seat_number')
    
    # Check for temporary locks
    now = timezone.now()
    
    seat_data = []
    for seat in seats:
        # Check if seat is temporarily locked
        lock_key = f"seat_lock_{seat.id}"
        locked_by = cache.get(lock_key)
        
        is_locked = locked_by is not None
        is_available = seat.is_available and not is_locked
        
        seat_data.append({
            'id': seat.id,
            'seat_number': seat.seat_number,
            'row_number': seat.row_number,
            'seat_class': seat.seat_class,
            'seat_class_display': seat.get_seat_class_display(),
            'position': seat.position,
            'position_display': seat.get_position_display(),
            'is_available': is_available,
            'is_locked': is_locked,
            'fare': float(seat.get_fare())
        })
    
    # Get layout configuration
    layout_config = trip.bus.seat_layout.layout_config
    
    return JsonResponse({
        'seats': seat_data,
        'layout': layout_config,
        'total_seats': trip.bus.seat_layout.total_seats,
        'total_rows': trip.bus.seat_layout.total_rows
    })


@require_http_methods(["POST"])
def api_lock_seat(request):
    """API to temporarily lock a seat (2 minutes)"""
    try:
        data = json.loads(request.body)
        seat_id = data.get('seat_id')
        session_id = request.session.session_key
        
        if not session_id:
            request.session.create()
            session_id = request.session.session_key
        
        seat = get_object_or_404(Seat, id=seat_id)
        
        if not seat.is_available:
            return JsonResponse({
                'success': False,
                'error': 'Seat is already booked'
            }, status=400)
        
        lock_key = f"seat_lock_{seat_id}"
        existing_lock = cache.get(lock_key)
        
        if existing_lock and existing_lock != session_id:
            return JsonResponse({
                'success': False,
                'error': 'Seat is currently being selected by another user'
            }, status=400)
        
        # Lock for 2 minutes
        cache.set(lock_key, session_id, 120)
        
        return JsonResponse({
            'success': True,
            'seat_id': seat_id,
            'expires_in': 120
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)


@require_http_methods(["POST"])
def api_unlock_seat(request):
    """API to unlock a temporarily locked seat"""
    try:
        data = json.loads(request.body)
        seat_id = data.get('seat_id')
        session_id = request.session.session_key
        
        lock_key = f"seat_lock_{seat_id}"
        existing_lock = cache.get(lock_key)
        
        if existing_lock == session_id:
            cache.delete(lock_key)
            return JsonResponse({'success': True})
        
        return JsonResponse({
            'success': False,
            'error': 'Seat not locked by this session'
        }, status=400)
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)


@require_http_methods(["GET"])
def api_get_boarding_points(request, trip_id):
    """API to get boarding and dropping points for a trip"""
    trip = get_object_or_404(Trip, id=trip_id)
    
    # Get route stops
    stops = trip.route.stops.select_related('boarding_point__location').all()
    
    boarding_points = []
    dropping_points = []
    
    for stop in stops:
        # IMPORTANT: Return the RouteStop ID, not the BoardingPoint ID
        point_data = {
            'id': stop.id,  # This is the RouteStop ID - what the booking expects
            'boarding_point_id': stop.boarding_point.id,  # This is the actual BoardingPoint ID
            'name': stop.boarding_point.name,
            'location': stop.boarding_point.location.name,
            'address': stop.boarding_point.address,
            'display': f"{stop.boarding_point.name} - {stop.boarding_point.location.name}"
        }
        
        if stop.is_pickup:
            boarding_points.append(point_data)
        if stop.is_dropoff:
            dropping_points.append(point_data)
    
    return JsonResponse({
        'boarding_points': boarding_points,
        'dropping_points': dropping_points
    })



@require_http_methods(["POST"])
def api_calculate_total(request):
    """API to calculate total booking amount"""
    try:
        data = json.loads(request.body)
        seat_ids = data.get('seat_ids', [])
        
        if not seat_ids:
            return JsonResponse({'error': 'No seats selected'}, status=400)
        
        seats = Seat.objects.filter(id__in=seat_ids)
        
        if seats.count() != len(seat_ids):
            return JsonResponse({'error': 'Some seats not found'}, status=400)
        
        total = sum(seat.get_fare() for seat in seats)
        
        seat_details = [
            {
                'seat_number': seat.seat_number,
                'seat_class': seat.get_seat_class_display(),
                'fare': float(seat.get_fare())
            }
            for seat in seats
        ]
        
        return JsonResponse({
            'success': True,
            'total': float(total),
            'seats': seat_details,
            'seat_count': len(seat_ids)
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)


import logging
from django.views.decorators.csrf import csrf_exempt

logger = logging.getLogger(__name__)

@csrf_exempt  # Add this if you're getting CSRF token errors
@require_http_methods(["POST"])
def api_create_booking(request):
    """API to create a booking with payment initiation"""
    try:
        # Log the raw request body for debugging
        logger.info(f"Request body: {request.body.decode('utf-8')}")
        
        data = json.loads(request.body)
        logger.info(f"Parsed data: {data}")
        
        # Validate required fields
        required_fields = [
            'trip_id', 'seat_ids', 'boarding_point_id', 
            'dropping_point_id', 'full_name', 'id_number', 
            'email', 'phone'
        ]
        
        missing_fields = []
        for field in required_fields:
            if field not in data or not data.get(field):
                missing_fields.append(field)
        
        if missing_fields:
            error_msg = f'Missing required fields: {", ".join(missing_fields)}'
            logger.error(error_msg)
            return JsonResponse({
                'error': error_msg,
                'missing_fields': missing_fields
            }, status=400)
        
        # Validate data types
        try:
            trip_id = int(data['trip_id'])
            boarding_point_id = int(data['boarding_point_id'])
            dropping_point_id = int(data['dropping_point_id'])
            seat_ids = [int(sid) for sid in data['seat_ids']]
        except (ValueError, TypeError) as e:
            logger.error(f"Invalid data type: {str(e)}")
            return JsonResponse({
                'error': 'Invalid data type for numeric fields'
            }, status=400)
        
        # Validate seat_ids is not empty
        if not seat_ids:
            return JsonResponse({
                'error': 'At least one seat must be selected'
            }, status=400)
        
        trip = get_object_or_404(Trip, id=trip_id)
        
        # Verify session has locks on these seats
        if not request.session.session_key:
            request.session.create()
        
        session_id = request.session.session_key
        logger.info(f"Session ID: {session_id}")
        
        locked_seats = []
        unlocked_seats = []
        
        for seat_id in seat_ids:
            lock_key = f"seat_lock_{seat_id}"
            locked_by = cache.get(lock_key)
            logger.info(f"Seat {seat_id} locked by: {locked_by}")
            
            if locked_by == session_id:
                locked_seats.append(seat_id)
            else:
                unlocked_seats.append(seat_id)
        
        if unlocked_seats:
            return JsonResponse({
                'error': f'Seat locks expired or invalid for seats: {unlocked_seats}. Please reselect seats.'
            }, status=400)
        
        # Get seats and verify availability
        seats = Seat.objects.filter(id__in=seat_ids, is_available=True)
        
        if seats.count() != len(seat_ids):
            unavailable = set(seat_ids) - set(seats.values_list('id', flat=True))
            return JsonResponse({
                'error': f'Some seats are no longer available: {list(unavailable)}'
            }, status=400)
        
        # Validate boarding and dropping points
        try:
            boarding_route_stop = RouteStop.objects.select_related('boarding_point').get(id=boarding_point_id)
            dropping_route_stop = RouteStop.objects.select_related('boarding_point').get(id=dropping_point_id)
            
            boarding_point = boarding_route_stop.boarding_point
            dropping_point = dropping_route_stop.boarding_point
            
        except RouteStop.DoesNotExist:
            return JsonResponse({
                'error': 'Invalid boarding or dropping point'
            }, status=400)
        
        # Calculate total
        total_amount = sum(seat.get_fare() for seat in seats)
        
        # Create booking
        booking = Booking.objects.create(
            trip=trip,
            customer_full_name=data['full_name'],
            customer_id_number=data['id_number'],
            customer_email=data['email'],
            customer_phone=data['phone'],
            boarding_point=boarding_point,
            dropping_point=dropping_point,
            total_amount=total_amount,
            status='pending'
        )
        
        logger.info(f"Created booking: {booking.booking_reference}")
        
        # Create seat bookings and mark seats as unavailable
        for seat in seats:
            SeatBooking.objects.create(
                booking=booking,
                seat=seat,
                fare=seat.get_fare()
            )
            seat.is_available = False
            seat.save()
            
            # Remove lock
            lock_key = f"seat_lock_{seat.id}"
            cache.delete(lock_key)
            logger.info(f"Removed lock for seat {seat.id}")
        
        # Create payment record
        payment = Payment.objects.create(
            booking=booking,
            transaction_id=f"TXN{booking.booking_reference}",
            payment_method='mpesa',
            amount=total_amount,
            mpesa_phone=data['phone'],
            status='initiated'
        )
        
        logger.info(f"Payment initiated: {payment.transaction_id}")
        
        # TODO: Initiate M-Pesa STK Push here
        
        return JsonResponse({
            'success': True,
            'booking_reference': booking.booking_reference,
            'total_amount': float(total_amount),
            'payment_id': payment.id,
            'message': 'Booking created. Please complete payment.'
        })
        
    except json.JSONDecodeError as e:
        logger.error(f"JSON decode error: {str(e)}")
        return JsonResponse({
            'error': 'Invalid JSON format',
            'details': str(e)
        }, status=400)
    except Exception as e:
        logger.exception(f"Unexpected error in booking creation: {str(e)}")
        return JsonResponse({
            'error': 'An unexpected error occurred',
            'details': str(e)
        }, status=500)



from django.shortcuts import render
from django.db.models import Sum, Count, Q, Avg
from django.utils import timezone
from datetime import timedelta, datetime
from decimal import Decimal
from .models import (
    Booking, Trip, Bus, Payment, Route, 
    SeatBooking, BusOperator, Review
)


def admin_dashboard(request):
    """
    Admin dashboard view with statistics, charts data, and recent activity
    """
    today = timezone.now().date()
    yesterday = today - timedelta(days=1)
    last_30_days = today - timedelta(days=30)
    last_7_days = today - timedelta(days=7)
    
    # ============== TODAY'S STATS ==============
    
    # Today's bookings count
    today_bookings = Booking.objects.filter(
        created_at__date=today
    ).count()
    
    # Yesterday's bookings for comparison
    yesterday_bookings = Booking.objects.filter(
        created_at__date=yesterday
    ).count()
    
    # Calculate percentage change
    if yesterday_bookings > 0:
        bookings_change = ((today_bookings - yesterday_bookings) / yesterday_bookings) * 100
    else:
        bookings_change = 100 if today_bookings > 0 else 0
    
    # Today's revenue
    today_revenue = Payment.objects.filter(
        created_at__date=today,
        status='completed'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Yesterday's revenue
    yesterday_revenue = Payment.objects.filter(
        created_at__date=yesterday,
        status='completed'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Revenue percentage change
    if yesterday_revenue > 0:
        revenue_change = ((today_revenue - yesterday_revenue) / yesterday_revenue) * 100
    else:
        revenue_change = 100 if today_revenue > 0 else 0
    
    # Active trips (scheduled or boarding today)
    active_trips = Trip.objects.filter(
        departure_date=today,
        status__in=['scheduled', 'boarding']
    ).count()
    
    # Cancelled trips today
    cancelled_today = Trip.objects.filter(
        departure_date=today,
        status='cancelled'
    ).count()
    
    # Pending payments
    pending_payments = Booking.objects.filter(
        status='pending'
    ).count()
    
    # Pending payments yesterday
    pending_yesterday = Booking.objects.filter(
        status='pending',
        created_at__date=yesterday
    ).count()
    
    # ============== REVENUE CHART DATA (Last 30 Days) ==============
    revenue_data = []
    revenue_labels = []
    
    for i in range(29, -1, -1):
        date = today - timedelta(days=i)
        daily_revenue = Payment.objects.filter(
            created_at__date=date,
            status='completed'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        revenue_data.append(float(daily_revenue))
        revenue_labels.append(date.strftime('%b %d'))
    
    # ============== BOOKING STATUS PIE CHART ==============
    booking_statuses = Booking.objects.filter(
        created_at__gte=last_30_days
    ).values('status').annotate(count=Count('id'))
    
    status_labels = []
    status_data = []
    status_colors = {
        'confirmed': '#3498db',
        'pending': '#f39c12',
        'cancelled': '#e74c3c',
        'completed': '#27ae60',
        'paid': '#9b59b6'
    }
    
    for status in booking_statuses:
        status_labels.append(status['status'].title())
        status_data.append(status['count'])
    
    # ============== TOP ROUTES BAR CHART ==============
    top_routes = Booking.objects.filter(
        created_at__gte=last_30_days
    ).values(
        'trip__route__origin__name',
        'trip__route__destination__name'
    ).annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    route_labels = []
    route_data = []
    
    for route in top_routes:
        origin = route['trip__route__origin__name']
        destination = route['trip__route__destination__name']
        route_labels.append(f"{origin} - {destination}")
        route_data.append(route['count'])
    
    # ============== PAYMENT METHODS DOUGHNUT CHART ==============
    payment_methods = Payment.objects.filter(
        created_at__gte=last_30_days,
        status='completed'
    ).values('payment_method').annotate(count=Count('id'))
    
    payment_labels = []
    payment_data = []
    payment_colors = {
        'mpesa': '#27ae60',
        'card': '#3498db',
        'cash': '#95a5a6'
    }
    
    for method in payment_methods:
        payment_labels.append(method['payment_method'].upper())
        payment_data.append(method['count'])
    
    # ============== OCCUPANCY RATE LINE CHART (Last 7 Days) ==============
    occupancy_data = []
    occupancy_labels = []
    
    for i in range(6, -1, -1):
        date = today - timedelta(days=i)
        
        # Get all trips for this date
        trips_on_date = Trip.objects.filter(
            departure_date=date,
            status__in=['departed', 'completed']
        )
        
        total_seats = 0
        booked_seats = 0
        
        for trip in trips_on_date:
            total_seats += trip.bus.seat_layout.total_seats
            booked = SeatBooking.objects.filter(
                seat__trip=trip,
                booking__status__in=['confirmed', 'paid', 'completed']
            ).count()
            booked_seats += booked
        
        # Calculate occupancy percentage
        if total_seats > 0:
            occupancy_rate = (booked_seats / total_seats) * 100
        else:
            occupancy_rate = 0
        
        occupancy_data.append(round(occupancy_rate, 1))
        occupancy_labels.append(date.strftime('%a'))
    
    # ============== RECENT BOOKINGS ==============
    recent_bookings = Booking.objects.select_related(
        'trip__route__origin',
        'trip__route__destination'
    ).order_by('-created_at')[:5]
    
    bookings_list = []
    for booking in recent_bookings:
        bookings_list.append({
            'reference': booking.booking_reference,
            'customer_name': booking.customer_full_name,
            'route': f"{booking.trip.route.origin.name} → {booking.trip.route.destination.name}",
            'amount': booking.total_amount,
            'status': booking.status,
            'created_at': booking.created_at
        })
    
    # ============== TODAY'S TRIPS ==============
    todays_trips = Trip.objects.filter(
        departure_date=today
    ).select_related(
        'bus',
        'route__origin',
        'route__destination'
    ).order_by('departure_time')[:5]
    
    trips_list = []
    for trip in todays_trips:
        total_seats = trip.bus.seat_layout.total_seats
        booked_seats = SeatBooking.objects.filter(
            seat__trip=trip,
            booking__status__in=['confirmed', 'paid', 'pending']
        ).count()
        
        occupancy_percentage = (booked_seats / total_seats * 100) if total_seats > 0 else 0
        
        # Determine occupancy level
        if occupancy_percentage >= 80:
            occupancy_level = 'high'
        elif occupancy_percentage >= 50:
            occupancy_level = 'medium'
        else:
            occupancy_level = 'low'
        
        trips_list.append({
            'time': trip.departure_time.strftime('%I:%M'),
            'period': trip.departure_time.strftime('%p'),
            'route': f"{trip.route.origin.name} → {trip.route.destination.name}",
            'bus_name': trip.bus.bus_name,
            'bus_type': trip.bus.get_bus_type_display(),
            'booked_seats': booked_seats,
            'total_seats': total_seats,
            'occupancy_percentage': round(occupancy_percentage, 0),
            'occupancy_level': occupancy_level,
            'status': trip.status
        })
    
    # ============== ADDITIONAL STATS ==============
    
    # Total buses
    total_buses = Bus.objects.filter(is_active=True).count()
    
    # Total operators
    total_operators = BusOperator.objects.filter(is_active=True).count()
    
    # Average rating
    avg_rating = Review.objects.aggregate(avg=Avg('rating'))['avg'] or 0
    
    # Total routes
    total_routes = Route.objects.filter(is_active=True).count()
    
    context = {
        # Stats
        'today_bookings': today_bookings,
        'bookings_change': round(bookings_change, 1),
        'bookings_change_positive': bookings_change >= 0,
        
        'today_revenue': today_revenue,
        'revenue_change': round(revenue_change, 1),
        'revenue_change_positive': revenue_change >= 0,
        
        'active_trips': active_trips,
        'cancelled_today': cancelled_today,
        
        'pending_payments': pending_payments,
        'pending_change': pending_yesterday - pending_payments,
        'pending_change_positive': (pending_yesterday - pending_payments) > 0,
        
        # Chart data - Revenue
        'revenue_labels': revenue_labels,
        'revenue_data': revenue_data,
        
        # Chart data - Booking Status
        'status_labels': status_labels,
        'status_data': status_data,
        
        # Chart data - Top Routes
        'route_labels': route_labels,
        'route_data': route_data,
        
        # Chart data - Payment Methods
        'payment_labels': payment_labels,
        'payment_data': payment_data,
        
        # Chart data - Occupancy
        'occupancy_labels': occupancy_labels,
        'occupancy_data': occupancy_data,
        
        # Recent data
        'recent_bookings': bookings_list,
        'todays_trips': trips_list,
        
        # Additional stats
        'total_buses': total_buses,
        'total_operators': total_operators,
        'avg_rating': round(avg_rating, 2),
        'total_routes': total_routes,
    }
    
    return render(request, 'admin/dashboard.html', context)



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q, Count, Avg
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from .models import Bus, BusOperator, SeatLayout, Amenity
from .forms import BusForm, BusOperatorForm, SeatLayoutForm
import json


# ============= BUS VIEWS =============

def bus_list(request):
    """Display all buses with filters"""
    buses = Bus.objects.select_related('operator', 'seat_layout').prefetch_related('amenities')
    
    # Filters
    search = request.GET.get('search', '')
    operator_id = request.GET.get('operator', '')
    bus_type = request.GET.get('type', '')
    status = request.GET.get('status', '')
    
    if search:
        buses = buses.filter(
            Q(bus_name__icontains=search) | 
            Q(registration_number__icontains=search)
        )
    
    if operator_id:
        buses = buses.filter(operator_id=operator_id)
    
    if bus_type:
        buses = buses.filter(bus_type=bus_type)
    
    if status == 'active':
        buses = buses.filter(is_active=True)
    elif status == 'inactive':
        buses = buses.filter(is_active=False)
    
    buses = buses.order_by('-created_at')
    
    # Get all operators for filter dropdown
    operators = BusOperator.objects.filter(is_active=True)
    
    context = {
        'buses': buses,
        'operators': operators,
        'search': search,
        'selected_operator': operator_id,
        'selected_type': bus_type,
        'selected_status': status,
        'bus_types': Bus.BUS_TYPE_CHOICES,
    }
    
    return render(request, 'buses/bus_list.html', context)


def bus_detail(request, pk):
    """Display detailed bus information with seat layout visualization"""
    bus = get_object_or_404(
        Bus.objects.select_related('operator', 'seat_layout').prefetch_related('amenities'),
        pk=pk
    )
    
    # Get seat layout configuration
    layout_config = bus.seat_layout.layout_config
    
    # Get upcoming trips count
    from django.utils import timezone
    upcoming_trips = bus.trips.filter(
        departure_date__gte=timezone.now().date(),
        is_active=True
    ).count()
    
    # Get recent reviews
    recent_reviews = bus.reviews.select_related('booking').order_by('-created_at')[:5]
    
    context = {
        'bus': bus,
        'layout_config': json.dumps(layout_config),
        'upcoming_trips': upcoming_trips,
        'recent_reviews': recent_reviews,
    }
    
    return render(request, 'buses/bus_detail.html', context)


def bus_form(request, pk=None):
    """Add or edit bus"""
    if pk:
        bus = get_object_or_404(Bus, pk=pk)
        title = f"Edit Bus: {bus.bus_name}"
    else:
        bus = None
        title = "Add New Bus"
    
    if request.method == 'POST':
        form = BusForm(request.POST, instance=bus)
        if form.is_valid():
            bus = form.save()
            messages.success(request, f"Bus '{bus.bus_name}' saved successfully!")
            return redirect('bus_detail', pk=bus.pk)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = BusForm(instance=bus)
    
    context = {
        'form': form,
        'title': title,
        'bus': bus,
    }
    
    return render(request, 'buses/bus_form.html', context)


@require_POST
def bus_delete(request, pk):
    """Delete bus"""
    bus = get_object_or_404(Bus, pk=pk)
    bus_name = bus.bus_name
    
    # Check if bus has any trips
    if bus.trips.exists():
        messages.error(request, f"Cannot delete '{bus_name}' because it has associated trips.")
    else:
        bus.delete()
        messages.success(request, f"Bus '{bus_name}' deleted successfully!")
    
    return redirect('bus_list')


@require_POST
def bus_toggle_status(request, pk):
    """Toggle bus active status"""
    bus = get_object_or_404(Bus, pk=pk)
    bus.is_active = not bus.is_active
    bus.save()
    
    status = "activated" if bus.is_active else "deactivated"
    messages.success(request, f"Bus '{bus.bus_name}' {status} successfully!")
    
    return redirect('bus_detail', pk=pk)


# ============= BUS OPERATOR VIEWS =============

def operator_list(request):
    """Display all bus operators"""
    operators = BusOperator.objects.annotate(
        bus_count=Count('buses')
    ).order_by('-created_at')
    
    search = request.GET.get('search', '')
    if search:
        operators = operators.filter(
            Q(name__icontains=search) | 
            Q(contact_email__icontains=search)
        )
    
    context = {
        'operators': operators,
        'search': search,
    }
    
    return render(request, 'buses/operator_list.html', context)


def operator_detail(request, pk):
    """Display operator details and their buses"""
    operator = get_object_or_404(BusOperator, pk=pk)
    buses = operator.buses.select_related('seat_layout').prefetch_related('amenities')
    
    context = {
        'operator': operator,
        'buses': buses,
    }
    
    return render(request, 'buses/operator_detail.html', context)


def operator_form(request, pk=None):
    """Add or edit operator"""
    if pk:
        operator = get_object_or_404(BusOperator, pk=pk)
        title = f"Edit Operator: {operator.name}"
    else:
        operator = None
        title = "Add New Operator"
    
    if request.method == 'POST':
        form = BusOperatorForm(request.POST, request.FILES, instance=operator)
        if form.is_valid():
            operator = form.save()
            messages.success(request, f"Operator '{operator.name}' saved successfully!")
            return redirect('operator_detail', pk=operator.pk)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = BusOperatorForm(instance=operator)
    
    context = {
        'form': form,
        'title': title,
        'operator': operator,
    }
    
    return render(request, 'buses/operator_form.html', context)


@require_POST
def operator_delete(request, pk):
    """Delete operator"""
    operator = get_object_or_404(BusOperator, pk=pk)
    operator_name = operator.name
    
    if operator.buses.exists():
        messages.error(request, f"Cannot delete '{operator_name}' because it has associated buses.")
    else:
        operator.delete()
        messages.success(request, f"Operator '{operator_name}' deleted successfully!")
    
    return redirect('operator_list')


# ============= SEAT LAYOUT VIEWS =============

def layout_list(request):
    """Display all seat layouts"""
    layouts = SeatLayout.objects.annotate(
        bus_count=Count('bus')
    ).order_by('-id')
    
    context = {
        'layouts': layouts,
    }
    
    return render(request, 'buses/layout_list.html', context)


def layout_detail(request, pk):
    """Display seat layout with visualization"""
    layout = get_object_or_404(SeatLayout, pk=pk)
    
    # Get buses using this layout
    buses = layout.bus_set.select_related('operator')
    
    context = {
        'layout': layout,
        'layout_config': json.dumps(layout.layout_config),
        'buses': buses,
    }
    
    return render(request, 'buses/layout_detail.html', context)


def layout_form(request, pk=None):
    """Add or edit seat layout"""
    if pk:
        layout = get_object_or_404(SeatLayout, pk=pk)
        title = f"Edit Layout: {layout.name}"
    else:
        layout = None
        title = "Add New Seat Layout"
    
    if request.method == 'POST':
        form = SeatLayoutForm(request.POST, request.FILES, instance=layout)
        if form.is_valid():
            layout = form.save()
            messages.success(request, f"Seat layout '{layout.name}' saved successfully!")
            return redirect('layout_detail', pk=layout.pk)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = SeatLayoutForm(instance=layout)
    
    context = {
        'form': form,
        'title': title,
        'layout': layout,
    }
    
    return render(request, 'buses/layout_form.html', context)


@require_POST
def layout_delete(request, pk):
    """Delete seat layout"""
    layout = get_object_or_404(SeatLayout, pk=pk)
    layout_name = layout.name
    
    if layout.bus_set.exists():
        messages.error(request, f"Cannot delete '{layout_name}' because it's being used by buses.")
    else:
        layout.delete()
        messages.success(request, f"Seat layout '{layout_name}' deleted successfully!")
    
    return redirect('layout_list')


# ============= AJAX/API VIEWS =============

def get_layout_preview(request, pk):
    """Get layout configuration for preview"""
    layout = get_object_or_404(SeatLayout, pk=pk)
    
    return JsonResponse({
        'success': True,
        'layout': layout.layout_config,
        'name': layout.name,
        'total_seats': layout.total_seats,
    })


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q, Count, Sum, Prefetch
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from datetime import datetime, timedelta
from .models import Trip, Bus, Route, Booking, SeatBooking, Seat
from .forms import TripForm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO


# ============= TRIP LIST VIEW =============
def trip_list(request):
    """Display all trips with filters"""
    trips = Trip.objects.select_related(
        'bus', 'bus__operator', 'route', 'route__origin', 'route__destination'
    ).annotate(
        bookings_count=Count('bookings'),
        seats_booked=Count('bookings')  # 👈 replaced seat_bookings with bookings
    )
    
    # Filters
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    route_id = request.GET.get('route', '')
    operator_id = request.GET.get('operator', '')
    
    # Search by route or bus
    if search:
        trips = trips.filter(
            Q(route__origin__name__icontains=search) |
            Q(route__destination__name__icontains=search) |
            Q(bus__bus_name__icontains=search) |
            Q(bus__registration_number__icontains=search)
        )
    
    # Filter by status
    if status:
        trips = trips.filter(status=status)
    
    # Filter by date range
    if date_from:
        trips = trips.filter(departure_date__gte=date_from)
    if date_to:
        trips = trips.filter(departure_date__lte=date_to)
    
    # Filter by route
    if route_id:
        trips = trips.filter(route_id=route_id)
    
    # Filter by operator
    if operator_id:
        trips = trips.filter(bus__operator_id=operator_id)
    
    # Separate upcoming and past trips
    today = timezone.now().date()
    upcoming_trips = trips.filter(departure_date__gte=today).order_by('departure_date', 'departure_time')
    past_trips = trips.filter(departure_date__lt=today).order_by('-departure_date', '-departure_time')
    
    # Get filter options
    from .models import Route, BusOperator
    routes = Route.objects.filter(is_active=True)
    operators = BusOperator.objects.filter(is_active=True)
    
    context = {
        'upcoming_trips': upcoming_trips,
        'past_trips': past_trips,
        'routes': routes,
        'operators': operators,
        'search': search,
        'selected_status': status,
        'date_from': date_from,
        'date_to': date_to,
        'selected_route': route_id,
        'selected_operator': operator_id,
        'trip_statuses': Trip.TRIP_STATUS_CHOICES,
    }
    
    return render(request, 'trips/trip_list.html', context)

# ============= TRIP DETAIL VIEW =============

def trip_detail(request, pk):
    """Display detailed trip information with passenger list"""
    trip = get_object_or_404(
        Trip.objects.select_related(
            'bus', 'bus__operator', 'bus__seat_layout',
            'route', 'route__origin', 'route__destination'
        ),
        pk=pk
    )
    
    # Get all bookings for this trip
    bookings = Booking.objects.filter(
        trip=trip
    ).select_related(
        'boarding_point', 'dropping_point'
    ).prefetch_related(
        Prefetch('seat_bookings', queryset=SeatBooking.objects.select_related('seat'))
    ).order_by('-created_at')
    
    # Get seat availability
    total_seats = trip.bus.seat_layout.total_seats
    booked_seats = SeatBooking.objects.filter(
        booking__trip=trip,
        booking__status__in=['pending', 'confirmed', 'paid']
    ).count()
    available_seats = total_seats - booked_seats
    
    # Calculate revenue
    total_revenue = sum(booking.total_amount for booking in bookings if booking.status in ['paid', 'confirmed'])
    pending_revenue = sum(booking.total_amount for booking in bookings if booking.status == 'pending')
    
    # Get seat layout with booking status
    seats = Seat.objects.filter(trip=trip).select_related('trip').prefetch_related(
        Prefetch('bookings', queryset=SeatBooking.objects.select_related('booking'))
    ).order_by('row_number', 'seat_number')
    
    # Group bookings by status
    confirmed_bookings = bookings.filter(status__in=['confirmed', 'paid'])
    pending_bookings = bookings.filter(status='pending')
    cancelled_bookings = bookings.filter(status='cancelled')
    
    context = {
        'trip': trip,
        'bookings': bookings,
        'confirmed_bookings': confirmed_bookings,
        'pending_bookings': pending_bookings,
        'cancelled_bookings': cancelled_bookings,
        'total_seats': total_seats,
        'booked_seats': booked_seats,
        'available_seats': available_seats,
        'total_revenue': total_revenue,
        'pending_revenue': pending_revenue,
        'seats': seats,
        'layout_config': trip.bus.seat_layout.layout_config,
    }
    
    return render(request, 'trips/trip_detail.html', context)


# ============= EXPORT PASSENGERS TO EXCEL =============

def export_passengers(request, pk):
    """Export trip passengers to Excel"""
    trip = get_object_or_404(
        Trip.objects.select_related(
            'bus', 'bus__operator', 'route', 'route__origin', 'route__destination'
        ),
        pk=pk
    )
    
    # Get all confirmed bookings
    bookings = Booking.objects.filter(
        trip=trip,
        status__in=['confirmed', 'paid']
    ).select_related(
        'boarding_point', 'dropping_point'
    ).prefetch_related(
        'seat_bookings__seat'
    ).order_by('customer_full_name')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Passenger List"
    
    # Styles
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Trip information
    ws['A1'] = f"Passenger List - {trip.route.origin.name} to {trip.route.destination.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')
    
    ws['A2'] = f"Bus: {trip.bus.bus_name} ({trip.bus.registration_number})"
    ws.merge_cells('A2:G2')
    
    ws['A3'] = f"Date: {trip.departure_date.strftime('%B %d, %Y')} | Departure: {trip.departure_time.strftime('%I:%M %p')}"
    ws.merge_cells('A3:G3')
    
    ws['A4'] = f"Operator: {trip.bus.operator.name}"
    ws.merge_cells('A4:G4')
    
    ws['A5'] = f"Total Passengers: {bookings.count()}"
    ws.merge_cells('A5:G5')
    
    # Headers
    headers = ['#', 'Passenger Name', 'ID Number', 'Phone Number', 'Seat(s)', 'Boarding Point', 'Dropping Point']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    row_num = 8
    for index, booking in enumerate(bookings, start=1):
        # Get seat numbers
        seat_numbers = ', '.join([sb.seat.seat_number for sb in booking.seat_bookings.all()])
        
        data = [
            index,
            booking.customer_full_name,
            booking.customer_id_number,
            booking.customer_phone,
            seat_numbers,
            booking.boarding_point.name,
            booking.dropping_point.name
        ]
        
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row_num, column=col)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        row_num += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"passengers_{trip.route.origin.name}_{trip.route.destination.name}_{trip.departure_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


# ============= SCHEDULE TRIP VIEW =============

def trip_form(request, pk=None):
    """Schedule a new trip or edit existing one"""
    if pk:
        trip = get_object_or_404(Trip, pk=pk)
        title = f"Edit Trip: {trip.route}"
    else:
        trip = None
        title = "Schedule New Trip"
    
    if request.method == 'POST':
        form = TripForm(request.POST, instance=trip)
        if form.is_valid():
            trip = form.save(commit=False)
            
            # If new trip, create seats
            if not pk:
                trip.save()
                create_trip_seats(trip)
            else:
                trip.save()
            
            messages.success(request, f"Trip scheduled successfully for {trip.departure_date}!")
            return redirect('trip_detail', pk=trip.pk)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = TripForm(instance=trip)
    
    context = {
        'form': form,
        'title': title,
        'trip': trip,
    }
    
    return render(request, 'trips/trip_form.html', context)


def create_trip_seats(trip):
    """Create seat instances for a trip based on bus layout"""
    layout_config = trip.bus.seat_layout.layout_config
    
    if not layout_config or 'rows' not in layout_config:
        return
    
    seats_to_create = []
    for row in layout_config['rows']:
        for seat in row['seats']:
            seats_to_create.append(
                Seat(
                    trip=trip,
                    seat_number=f"{row['row']}{seat['position']}",
                    row_number=row['row'],
                    seat_class=seat.get('class', 'normal'),
                    position=seat.get('type', 'window'),
                    is_available=True
                )
            )
    
    Seat.objects.bulk_create(seats_to_create)


# ============= TRIP HISTORY VIEW =============

def trip_history(request):
    """Display completed and cancelled trips"""
    trips = Trip.objects.select_related(
        'bus', 'bus__operator', 'route', 'route__origin', 'route__destination'
    ).annotate(
        bookings_count=Count('bookings'),
        total_revenue=Sum('bookings__total_amount')
    ).filter(
        Q(status='completed') | Q(status='cancelled') | Q(departure_date__lt=timezone.now().date())
    ).order_by('-departure_date', '-departure_time')
    
    # Filters
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    month = request.GET.get('month', '')
    year = request.GET.get('year', '')
    
    if search:
        trips = trips.filter(
            Q(route__origin__name__icontains=search) |
            Q(route__destination__name__icontains=search) |
            Q(bus__bus_name__icontains=search)
        )
    
    if status:
        trips = trips.filter(status=status)
    
    if month:
        trips = trips.filter(departure_date__month=month)
    
    if year:
        trips = trips.filter(departure_date__year=year)
    
    # Calculate statistics
    total_trips = trips.count()
    total_revenue = sum(trip.total_revenue or 0 for trip in trips)
    completed_trips = trips.filter(status='completed').count()
    cancelled_trips = trips.filter(status='cancelled').count()
    
    context = {
        'trips': trips,
        'search': search,
        'selected_status': status,
        'selected_month': month,
        'selected_year': year,
        'total_trips': total_trips,
        'total_revenue': total_revenue,
        'completed_trips': completed_trips,
        'cancelled_trips': cancelled_trips,
    }
    
    return render(request, 'trips/trip_history.html', context)


# ============= AJAX/API VIEWS =============

@require_POST
def update_trip_status(request, pk):
    """Update trip status"""
    trip = get_object_or_404(Trip, pk=pk)
    new_status = request.POST.get('status')
    
    if new_status in dict(Trip.TRIP_STATUS_CHOICES):
        trip.status = new_status
        trip.save()
        messages.success(request, f"Trip status updated to {trip.get_status_display()}")
    else:
        messages.error(request, "Invalid status")
    
    return redirect('trip_detail', pk=pk)


@require_POST
def cancel_trip(request, pk):
    """Cancel a trip"""
    trip = get_object_or_404(Trip, pk=pk)
    
    if trip.status in ['completed', 'cancelled']:
        messages.error(request, "Cannot cancel this trip")
        return redirect('trip_detail', pk=pk)
    
    # Cancel all bookings
    bookings = Booking.objects.filter(trip=trip, status__in=['pending', 'confirmed'])
    cancelled_count = bookings.count()
    bookings.update(status='cancelled')
    
    # Update trip status
    trip.status = 'cancelled'
    trip.save()
    
    messages.success(request, f"Trip cancelled. {cancelled_count} booking(s) were cancelled.")
    return redirect('trip_detail', pk=pk)


def get_route_details(request, route_id):
    """Get route details for AJAX"""
    from .models import Route
    route = get_object_or_404(Route, pk=route_id)
    
    return JsonResponse({
        'origin': route.origin.name,
        'destination': route.destination.name,
        'distance': str(route.distance_km),
        'duration': str(route.estimated_duration),
    })



from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Q, Count, Sum, Prefetch
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .models import (
    Booking, SeatBooking, Payment, Trip, BusOperator, 
    Location, BoardingPoint
)


@staff_member_required
def booking_list(request):
    """View for all bookings with filters and search"""
    
    # Get filter parameters
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    payment_status = request.GET.get('payment_status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    operator = request.GET.get('operator', '')
    route = request.GET.get('route', '')
    
    # Base queryset with related data
    bookings = Booking.objects.select_related(
        'trip__bus__operator',
        'trip__route__origin',
        'trip__route__destination',
        'boarding_point__location',
        'dropping_point__location'
    ).prefetch_related(
        'seat_bookings__seat',
        'payments'
    ).annotate(
        seats_count=Count('seat_bookings')
    ).order_by('-created_at')
    
    # Apply search filter
    if search:
        bookings = bookings.filter(
            Q(booking_reference__icontains=search) |
            Q(customer_full_name__icontains=search) |
            Q(customer_phone__icontains=search) |
            Q(customer_email__icontains=search) |
            Q(customer_id_number__icontains=search)
        )
    
    # Apply status filter
    if status:
        bookings = bookings.filter(status=status)
    
    # Apply payment status filter (check latest payment)
    if payment_status:
        bookings = bookings.filter(payments__status=payment_status).distinct()
    
    # Apply date range filter
    if date_from:
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        bookings = bookings.filter(trip__departure_date__gte=date_from_obj)
    
    if date_to:
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        bookings = bookings.filter(trip__departure_date__lte=date_to_obj)
    
    # Apply operator filter
    if operator:
        bookings = bookings.filter(trip__bus__operator_id=operator)
    
    # Apply route filter
    if route:
        origin_id, destination_id = route.split('-')
        bookings = bookings.filter(
            trip__route__origin_id=origin_id,
            trip__route__destination_id=destination_id
        )
    
    # Get counts for badges
    all_count = bookings.count()
    pending_payments_count = Booking.objects.filter(
        status='pending'
    ).count()
    confirmed_count = bookings.filter(status__in=['paid', 'confirmed']).count()
    cancelled_count = bookings.filter(status='cancelled').count()
    
    # Get operators for filter
    operators = BusOperator.objects.filter(is_active=True)
    
    # Get unique routes for filter
    routes = Trip.objects.select_related(
        'route__origin', 'route__destination'
    ).values(
        'route__origin_id',
        'route__origin__name',
        'route__destination_id',
        'route__destination__name'
    ).distinct()
    
    # Booking status choices
    booking_statuses = Booking.BOOKING_STATUS_CHOICES
    payment_statuses = Payment.PAYMENT_STATUS_CHOICES
    
    context = {
        'bookings': bookings,
        'search': search,
        'selected_status': status,
        'selected_payment_status': payment_status,
        'date_from': date_from,
        'date_to': date_to,
        'selected_operator': operator,
        'selected_route': route,
        'all_count': all_count,
        'pending_payments_count': pending_payments_count,
        'confirmed_count': confirmed_count,
        'cancelled_count': cancelled_count,
        'operators': operators,
        'routes': routes,
        'booking_statuses': booking_statuses,
        'payment_statuses': payment_statuses,
    }
    
    return render(request, 'admin/bookings/booking_list.html', context)


@staff_member_required
def booking_detail(request, booking_id):
    """View for booking details"""
    
    booking = get_object_or_404(
        Booking.objects.select_related(
            'trip__bus__operator',
            'trip__bus__seat_layout',
            'trip__route__origin',
            'trip__route__destination',
            'boarding_point__location',
            'dropping_point__location'
        ).prefetch_related(
            'seat_bookings__seat',
            'payments',
            'trip__bus__amenities'
        ),
        pk=booking_id
    )
    
    # Get seat bookings with details
    seat_bookings = booking.seat_bookings.all()
    
    # Get payment history
    payments = booking.payments.order_by('-created_at')
    
    # Calculate payment summary
    total_paid = payments.filter(status='completed').aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    balance = booking.total_amount - total_paid
    
    context = {
        'booking': booking,
        'seat_bookings': seat_bookings,
        'payments': payments,
        'total_paid': total_paid,
        'balance': balance,
    }
    
    return render(request, 'admin/bookings/booking_detail.html', context)


@staff_member_required
def pending_payments(request):
    """View for bookings with pending payments"""
    
    # Get filter parameters
    search = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Base queryset
    bookings = Booking.objects.filter(
        status='pending'
    ).select_related(
        'trip__bus__operator',
        'trip__route__origin',
        'trip__route__destination'
    ).prefetch_related(
        'seat_bookings',
        'payments'
    ).annotate(
        seats_count=Count('seat_bookings')
    ).order_by('-created_at')
    
    # Apply search filter
    if search:
        bookings = bookings.filter(
            Q(booking_reference__icontains=search) |
            Q(customer_full_name__icontains=search) |
            Q(customer_phone__icontains=search)
        )
    
    # Apply date range filter
    if date_from:
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        bookings = bookings.filter(created_at__date__gte=date_from_obj)
    
    if date_to:
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        bookings = bookings.filter(created_at__date__lte=date_to_obj)
    
    context = {
        'bookings': bookings,
        'search': search,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'admin/bookings/pending_payments.html', context)


@staff_member_required
def payment_list(request):
    """View for all payment transactions"""
    
    # Get filter parameters
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    method = request.GET.get('method', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Base queryset
    payments = Payment.objects.select_related(
        'booking__trip__route__origin',
        'booking__trip__route__destination'
    ).order_by('-created_at')
    
    # Apply search filter
    if search:
        payments = payments.filter(
            Q(transaction_id__icontains=search) |
            Q(booking__booking_reference__icontains=search) |
            Q(booking__customer_full_name__icontains=search) |
            Q(mpesa_phone__icontains=search) |
            Q(mpesa_receipt__icontains=search)
        )
    
    # Apply status filter
    if status:
        payments = payments.filter(status=status)
    
    # Apply payment method filter
    if method:
        payments = payments.filter(payment_method=method)
    
    # Apply date range filter
    if date_from:
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        payments = payments.filter(created_at__date__gte=date_from_obj)
    
    if date_to:
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        payments = payments.filter(created_at__date__lte=date_to_obj)
    
    # Calculate summary statistics
    total_payments = payments.filter(status='completed').aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    pending_amount = payments.filter(status='pending').aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    failed_count = payments.filter(status='failed').count()
    
    context = {
        'payments': payments,
        'search': search,
        'selected_status': status,
        'selected_method': method,
        'date_from': date_from,
        'date_to': date_to,
        'payment_statuses': Payment.PAYMENT_STATUS_CHOICES,
        'payment_methods': Payment.PAYMENT_METHOD_CHOICES,
        'total_payments': total_payments,
        'pending_amount': pending_amount,
        'failed_count': failed_count,
    }
    
    return render(request, 'admin/bookings/payment_list.html', context)


@staff_member_required
def export_bookings(request):
    """Export bookings to Excel"""
    
    # Get filter parameters (same as booking_list)
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    payment_status = request.GET.get('payment_status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    operator = request.GET.get('operator', '')
    
    # Apply same filters as booking_list
    bookings = Booking.objects.select_related(
        'trip__bus__operator',
        'trip__route__origin',
        'trip__route__destination',
        'boarding_point__location',
        'dropping_point__location'
    ).prefetch_related(
        'seat_bookings__seat',
        'payments'
    ).order_by('-created_at')
    
    if search:
        bookings = bookings.filter(
            Q(booking_reference__icontains=search) |
            Q(customer_full_name__icontains=search) |
            Q(customer_phone__icontains=search) |
            Q(customer_email__icontains=search)
        )
    
    if status:
        bookings = bookings.filter(status=status)
    
    if payment_status:
        bookings = bookings.filter(payments__status=payment_status).distinct()
    
    if date_from:
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        bookings = bookings.filter(trip__departure_date__gte=date_from_obj)
    
    if date_to:
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        bookings = bookings.filter(trip__departure_date__lte=date_to_obj)
    
    if operator:
        bookings = bookings.filter(trip__bus__operator_id=operator)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bookings"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Booking Ref',
        'Customer Name',
        'ID Number',
        'Phone',
        'Email',
        'Route',
        'Travel Date',
        'Departure Time',
        'Bus',
        'Operator',
        'Boarding Point',
        'Dropping Point',
        'Seats',
        'Total Amount',
        'Status',
        'Payment Status',
        'Booking Date'
    ]
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data
    for row_idx, booking in enumerate(bookings, start=2):
        # Get seat numbers
        seat_numbers = ', '.join([
            sb.seat.seat_number for sb in booking.seat_bookings.all()
        ])
        
        # Get latest payment status
        latest_payment = booking.payments.order_by('-created_at').first()
        payment_status_display = latest_payment.get_status_display() if latest_payment else 'No Payment'
        
        data = [
            booking.booking_reference,
            booking.customer_full_name,
            booking.customer_id_number,
            booking.customer_phone,
            booking.customer_email,
            f"{booking.trip.route.origin.name} → {booking.trip.route.destination.name}",
            booking.trip.departure_date.strftime('%Y-%m-%d'),
            booking.trip.departure_time.strftime('%H:%M'),
            booking.trip.bus.bus_name,
            booking.trip.bus.operator.name,
            f"{booking.boarding_point.location.name} - {booking.boarding_point.name}",
            f"{booking.dropping_point.location.name} - {booking.dropping_point.name}",
            seat_numbers,
            float(booking.total_amount),
            booking.get_status_display(),
            payment_status_display,
            booking.created_at.strftime('%Y-%m-%d %H:%M')
        ]
        
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    
    # Adjust column widths
    column_widths = [15, 25, 15, 15, 30, 30, 12, 12, 20, 20, 30, 30, 15, 12, 12, 15, 18]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Freeze header row
    ws.freeze_panes = ws['A2']
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=bookings_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


@staff_member_required
def export_payments(request):
    """Export payments to Excel"""
    
    # Get filter parameters
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    method = request.GET.get('method', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Apply filters
    payments = Payment.objects.select_related(
        'booking__trip__route__origin',
        'booking__trip__route__destination'
    ).order_by('-created_at')
    
    if search:
        payments = payments.filter(
            Q(transaction_id__icontains=search) |
            Q(booking__booking_reference__icontains=search) |
            Q(mpesa_receipt__icontains=search)
        )
    
    if status:
        payments = payments.filter(status=status)
    
    if method:
        payments = payments.filter(payment_method=method)
    
    if date_from:
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        payments = payments.filter(created_at__date__gte=date_from_obj)
    
    if date_to:
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        payments = payments.filter(created_at__date__lte=date_to_obj)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Transaction ID',
        'Booking Ref',
        'Customer Name',
        'Payment Method',
        'Amount',
        'Status',
        'M-Pesa Phone',
        'M-Pesa Receipt',
        'Route',
        'Travel Date',
        'Payment Date'
    ]
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data
    for row_idx, payment in enumerate(payments, start=2):
        data = [
            payment.transaction_id,
            payment.booking.booking_reference,
            payment.booking.customer_full_name,
            payment.get_payment_method_display(),
            float(payment.amount),
            payment.get_status_display(),
            payment.mpesa_phone or 'N/A',
            payment.mpesa_receipt or 'N/A',
            f"{payment.booking.trip.route.origin.name} → {payment.booking.trip.route.destination.name}",
            payment.booking.trip.departure_date.strftime('%Y-%m-%d'),
            payment.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ]
        
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    
    # Adjust column widths
    column_widths = [20, 15, 25, 15, 12, 12, 15, 20, 30, 12, 20]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Freeze header row
    ws.freeze_panes = ws['A2']
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=payments_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


@staff_member_required
def cancel_booking(request, booking_id):
    """Cancel a booking"""
    booking = get_object_or_404(Booking, pk=booking_id)
    
    if request.method == 'POST':
        if booking.status not in ['cancelled', 'completed']:
            booking.status = 'cancelled'
            booking.save()
            
            # Update seat availability
            for seat_booking in booking.seat_bookings.all():
                seat_booking.seat.is_available = True
                seat_booking.seat.save()
        
        return redirect('booking_detail', booking_id=booking.id)
    
    return redirect('booking_detail', booking_id=booking.id)


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Q, Count, Prefetch
from django.utils import timezone
from .models import Route, Location, BoardingPoint, RouteStop, Trip


# ============================================
# ROUTES VIEWS
# ============================================

def route_list(request):
    """Display all routes with filtering"""
    routes = Route.objects.select_related(
        'origin', 'destination'
    ).annotate(
        trips_count=Count('trips'),
        stops_count=Count('stops')
    ).filter(is_active=True)
    
    # Search filter
    search = request.GET.get('search', '')
    if search:
        routes = routes.filter(
            Q(origin__name__icontains=search) |
            Q(destination__name__icontains=search)
        )
    
    # Origin filter
    origin_id = request.GET.get('origin', '')
    if origin_id:
        routes = routes.filter(origin_id=origin_id)
    
    # Destination filter
    destination_id = request.GET.get('destination', '')
    if destination_id:
        routes = routes.filter(destination_id=destination_id)
    
    # Get all active locations for filters
    locations = Location.objects.filter(is_active=True).order_by('name')
    
    context = {
        'routes': routes.order_by('origin__name', 'destination__name'),
        'locations': locations,
        'search': search,
        'selected_origin': origin_id,
        'selected_destination': destination_id,
    }
    
    return render(request, 'admin/routes/route_list.html', context)


def route_detail(request, pk):
    """Display detailed route information including boarding points and stops"""
    route = get_object_or_404(
        Route.objects.select_related('origin', 'destination'),
        pk=pk
    )
    
    # Get route stops with boarding points
    route_stops = RouteStop.objects.filter(route=route).select_related(
        'boarding_point', 'boarding_point__location'
    ).order_by('stop_order')
    
    # Get upcoming trips for this route
    upcoming_trips = Trip.objects.filter(
        route=route,
        departure_date__gte=timezone.now().date(),
        is_active=True
    ).select_related('bus', 'bus__operator').order_by(
        'departure_date', 'departure_time'
    )[:5]
    
    # Calculate statistics
    total_trips = Trip.objects.filter(route=route).count()
    active_trips = upcoming_trips.count()
    
    # Separate pickup and dropoff points
    pickup_points = route_stops.filter(is_pickup=True)
    dropoff_points = route_stops.filter(is_dropoff=True)
    
    context = {
        'route': route,
        'route_stops': route_stops,
        'pickup_points': pickup_points,
        'dropoff_points': dropoff_points,
        'upcoming_trips': upcoming_trips,
        'total_trips': total_trips,
        'active_trips': active_trips,
    }
    
    return render(request, 'admin/routes/route_detail.html', context)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.http import JsonResponse
from datetime import timedelta
import json
from .models import Route, RouteStop, BoardingPoint, Location


def route_edit(request, pk):
    """Edit route with dynamic stops management"""
    route = get_object_or_404(Route, pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Update basic route information
                route.distance_km = request.POST.get('distance_km')
                route.is_active = request.POST.get('is_active') == 'on'
                
                # Convert duration hours to timedelta
                duration_hours = float(request.POST.get('duration_hours', 0))
                route.estimated_duration = timedelta(hours=duration_hours)
                route.save()
                
                # Delete existing stops
                route.stops.all().delete()
                
                # Process stops from form data
                stops_data = {}
                for key, value in request.POST.items():
                    if key.startswith('stops['):
                        # Parse: stops[1][boarding_point]
                        parts = key.replace('stops[', '').replace(']', '').split('[')
                        stop_id = parts[0]
                        field_name = parts[1] if len(parts) > 1 else None
                        
                        if stop_id not in stops_data:
                            stops_data[stop_id] = {}
                        
                        if field_name:
                            stops_data[stop_id][field_name] = value
                
                # Create new stops
                for stop_id, stop_data in stops_data.items():
                    if not stop_data.get('boarding_point'):
                        continue
                    
                    # Parse time_from_origin (HH:MM format)
                    time_str = stop_data.get('time_from_origin', '00:00')
                    try:
                        hours, minutes = map(int, time_str.split(':'))
                        time_from_origin = timedelta(hours=hours, minutes=minutes)
                    except:
                        time_from_origin = timedelta(0)
                    
                    RouteStop.objects.create(
                        route=route,
                        boarding_point_id=stop_data.get('boarding_point'),
                        stop_order=int(stop_data.get('stop_order', 0)),
                        time_from_origin=time_from_origin,
                        is_pickup=stop_data.get('is_pickup') == 'on' or stop_data.get('is_origin') == 'true',
                        is_dropoff=stop_data.get('is_dropoff') == 'on' or stop_data.get('is_destination') == 'true'
                    )
                
                messages.success(request, f'Route "{route}" updated successfully!')
                return redirect('route_detail', pk=route.pk)
                
        except Exception as e:
            messages.error(request, f'Error updating route: {str(e)}')
    
    # GET request - prepare data for template
    existing_stops = []
    for stop in route.stops.all():
        # Convert timedelta to HH:MM format
        total_seconds = int(stop.time_from_origin.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        time_str = f"{hours:02d}:{minutes:02d}"
        
        # Determine if origin or destination
        is_origin = stop.stop_order == 1
        is_destination = stop == route.stops.last()
        
        existing_stops.append({
            'id': stop.id,
            'boarding_point_id': stop.boarding_point.id,
            'boarding_point_name': stop.boarding_point.name,
            'stop_order': stop.stop_order,
            'time_from_origin': time_str,
            'is_pickup': stop.is_pickup,
            'is_dropoff': stop.is_dropoff,
            'is_origin': is_origin,
            'is_destination': is_destination,
            'is_food_stop': False,  # Add this field to your model if needed
            'break_duration': 30
        })
    
    # Get all boarding points for the dropdowns
    boarding_points = []
    for point in BoardingPoint.objects.filter(is_active=True).select_related('location'):
        boarding_points.append({
            'id': point.id,
            'name': point.name,
            'location_name': point.location.name,
            'location_id': point.location.id
        })
    
    # Calculate duration in hours for the form
    duration_hours = route.estimated_duration.total_seconds() / 3600
    
    context = {
        'route': route,
        'existing_stops': json.dumps(existing_stops),
        'boarding_points_json': json.dumps(boarding_points),
        'all_boarding_points': BoardingPoint.objects.filter(is_active=True).select_related('location'),
        'all_locations': Location.objects.filter(is_active=True),
        'duration_hours': duration_hours
    }
    
    return render(request, 'admin/routes/route_edit.html', context)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.http import JsonResponse
from datetime import timedelta
import json
from .models import Route, RouteStop, BoardingPoint, Location


def route_create(request):
    """Create new route with stops"""
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Get origin and destination
                origin_id = request.POST.get('origin')
                destination_id = request.POST.get('destination')
                
                if not origin_id or not destination_id:
                    messages.error(request, 'Please select both origin and destination.')
                    return redirect('route_create')
                
                if origin_id == destination_id:
                    messages.error(request, 'Origin and destination must be different.')
                    return redirect('route_create')
                
                # Check if route already exists
                existing_route = Route.objects.filter(
                    origin_id=origin_id,
                    destination_id=destination_id
                ).first()
                
                if existing_route:
                    messages.warning(
                        request, 
                        f'Route from {existing_route.origin.name} to {existing_route.destination.name} already exists. Redirecting to edit.'
                    )
                    return redirect('route_edit', pk=existing_route.pk)
                
                # Create the route
                route = Route.objects.create(
                    origin_id=origin_id,
                    destination_id=destination_id,
                    distance_km=request.POST.get('distance_km', 0),
                    estimated_duration=timedelta(hours=float(request.POST.get('duration_hours', 0))),
                    is_active=request.POST.get('is_active') == 'on'
                )
                
                # Process stops from form data
                stops_data = {}
                for key, value in request.POST.items():
                    if key.startswith('stops['):
                        # Parse: stops[1][boarding_point]
                        parts = key.replace('stops[', '').replace(']', '').split('[')
                        stop_id = parts[0]
                        field_name = parts[1] if len(parts) > 1 else None
                        
                        if stop_id not in stops_data:
                            stops_data[stop_id] = {}
                        
                        if field_name:
                            stops_data[stop_id][field_name] = value
                
                # Create stops
                created_stops = 0
                for stop_id, stop_data in stops_data.items():
                    boarding_point_id = stop_data.get('boarding_point')
                    if not boarding_point_id:
                        continue
                    
                    # Parse time_from_origin (HH:MM format)
                    time_str = stop_data.get('time_from_origin', '00:00')
                    try:
                        hours, minutes = map(int, time_str.split(':'))
                        time_from_origin = timedelta(hours=hours, minutes=minutes)
                    except:
                        time_from_origin = timedelta(0)
                    
                    # Parse break_duration if it's a food stop
                    break_duration = None
                    if stop_data.get('is_food_stop') == 'on':
                        try:
                            break_minutes = int(stop_data.get('break_duration', 30))
                            break_duration = timedelta(minutes=break_minutes)
                        except:
                            break_duration = timedelta(minutes=30)
                    
                    RouteStop.objects.create(
                        route=route,
                        boarding_point_id=boarding_point_id,
                        stop_order=int(stop_data.get('stop_order', 0)),
                        time_from_origin=time_from_origin,
                        is_pickup=stop_data.get('is_pickup') == 'on' or stop_data.get('is_origin') == 'true',
                        is_dropoff=stop_data.get('is_dropoff') == 'on' or stop_data.get('is_destination') == 'true',
                        # Optional fields if you have the enhanced model
                        # is_food_stop=stop_data.get('is_food_stop') == 'on',
                        # break_duration=break_duration,
                        # notes=stop_data.get('notes', '')
                    )
                    created_stops += 1
                
                if created_stops < 2:
                    messages.warning(
                        request, 
                        f'Route created but only {created_stops} stop(s) added. Please add more stops.'
                    )
                    return redirect('route_edit', pk=route.pk)
                
                messages.success(
                    request, 
                    f'Route "{route.origin.name} → {route.destination.name}" created successfully with {created_stops} stops!'
                )
                return redirect('route_detail', pk=route.pk)
                
        except Exception as e:
            messages.error(request, f'Error creating route: {str(e)}')
            return redirect('route_create')
    
    # GET request - prepare data for template
    # Get all boarding points for the dropdowns
    boarding_points = []
    for point in BoardingPoint.objects.filter(is_active=True).select_related('location'):
        boarding_points.append({
            'id': point.id,
            'name': point.name,
            'location_name': point.location.name,
            'location_id': point.location.id,
            'address': point.address,
            'landmark': point.landmark or ''
        })
    
    context = {
        'boarding_points_json': json.dumps(boarding_points),
        'all_boarding_points': BoardingPoint.objects.filter(is_active=True).select_related('location'),
        'all_locations': Location.objects.filter(is_active=True).order_by('name'),
    }
    
    return render(request, 'admin/routes/route_create.html', context)


# Helper function to calculate distance between locations (optional)
def calculate_route_distance(origin_id, destination_id):
    """
    Calculate estimated distance between two locations
    You can integrate with Google Maps API or use a distance matrix
    For now, returns a default value
    """
    # TODO: Integrate with mapping service
    return 100.0  # Default 100km


# Helper function to estimate duration (optional)
def estimate_route_duration(distance_km):
    """
    Estimate travel duration based on distance
    Assumes average speed of 60 km/h
    """
    average_speed = 60  # km/h
    hours = distance_km / average_speed
    return timedelta(hours=hours)


# Optional: AJAX endpoint to get boarding points by location
def get_boarding_points_by_location(request, location_id):
    """API endpoint to fetch boarding points for a specific location"""
    points = BoardingPoint.objects.filter(
        location_id=location_id, 
        is_active=True
    ).values('id', 'name', 'address', 'landmark')
    
    return JsonResponse({
        'success': True,
        'points': list(points)
    })


# Optional: Validate route stops
def validate_route_stops(request, pk):
    """Validate that route stops are properly configured"""
    route = get_object_or_404(Route, pk=pk)
    stops = route.stops.all().order_by('stop_order')
    
    errors = []
    warnings = []
    
    if stops.count() < 2:
        errors.append("Route must have at least 2 stops (origin and destination)")
    
    if stops.exists():
        # Check first stop is pickup only
        first_stop = stops.first()
        if not first_stop.is_pickup or first_stop.is_dropoff:
            warnings.append("First stop should typically be pickup only")
        
        # Check last stop is dropoff only
        last_stop = stops.last()
        if not last_stop.is_dropoff or last_stop.is_pickup:
            warnings.append("Last stop should typically be dropoff only")
        
        # Check time progression
        prev_time = timedelta(0)
        for stop in stops:
            if stop.time_from_origin < prev_time:
                errors.append(f"Stop {stop.stop_order} has invalid time progression")
            prev_time = stop.time_from_origin
    
    return JsonResponse({
        'valid': len(errors) == 0,
        'errors': errors,
        'warnings': warnings
    })


# ============================================
# LOCATIONS VIEWS
# ============================================

def location_list(request):
    """Display all locations"""
    locations = Location.objects.annotate(
        boarding_points_count=Count('boarding_points'),
        routes_from_count=Count('routes_from', filter=Q(routes_from__is_active=True)),
        routes_to_count=Count('routes_to', filter=Q(routes_to__is_active=True))
    ).filter(is_active=True)
    
    # Search filter
    search = request.GET.get('search', '')
    if search:
        locations = locations.filter(
            Q(name__icontains=search) |
            Q(county__icontains=search)
        )
    
    # County filter
    county = request.GET.get('county', '')
    if county:
        locations = locations.filter(county=county)
    
    # Get unique counties for filter
    counties = Location.objects.filter(
        is_active=True
    ).values_list('county', flat=True).distinct().order_by('county')
    
    context = {
        'locations': locations.order_by('name'),
        'counties': counties,
        'search': search,
        'selected_county': county,
    }
    
    return render(request, 'admin/locations/location_list.html', context)


def location_detail(request, pk):
    """Display detailed location information"""
    location = get_object_or_404(Location, pk=pk)
    
    # Get boarding points for this location
    boarding_points = BoardingPoint.objects.filter(
        location=location,
        is_active=True
    ).order_by('name')
    
    # Get routes from this location
    routes_from = Route.objects.filter(
        origin=location,
        is_active=True
    ).select_related('destination').annotate(
        trips_count=Count('trips')
    )
    
    # Get routes to this location
    routes_to = Route.objects.filter(
        destination=location,
        is_active=True
    ).select_related('origin').annotate(
        trips_count=Count('trips')
    )
    
    context = {
        'location': location,
        'boarding_points': boarding_points,
        'routes_from': routes_from,
        'routes_to': routes_to,
    }
    
    return render(request, 'admin/locations/location_detail.html', context)


# ============================================
# BOARDING POINTS VIEWS
# ============================================

def boarding_point_list(request):
    """Display all boarding points"""
    boarding_points = BoardingPoint.objects.select_related(
        'location'
    ).filter(is_active=True)
    
    # Search filter
    search = request.GET.get('search', '')
    if search:
        boarding_points = boarding_points.filter(
            Q(name__icontains=search) |
            Q(location__name__icontains=search) |
            Q(address__icontains=search) |
            Q(landmark__icontains=search)
        )
    
    # Location filter
    location_id = request.GET.get('location', '')
    if location_id:
        boarding_points = boarding_points.filter(location_id=location_id)
    
    # Get all locations for filter
    locations = Location.objects.filter(is_active=True).order_by('name')
    
    context = {
        'boarding_points': boarding_points.order_by('location__name', 'name'),
        'locations': locations,
        'search': search,
        'selected_location': location_id,
    }
    
    return render(request, 'admin/boarding_points/boarding_point_list.html', context)


def boarding_point_detail(request, pk):
    """Display detailed boarding point information"""
    boarding_point = get_object_or_404(
        BoardingPoint.objects.select_related('location'),
        pk=pk
    )
    
    # Get routes that include this boarding point
    route_stops = RouteStop.objects.filter(
        boarding_point=boarding_point
    ).select_related(
        'route', 'route__origin', 'route__destination'
    ).order_by('route__origin__name')
    
    # Separate pickup and dropoff routes
    pickup_routes = route_stops.filter(is_pickup=True)
    dropoff_routes = route_stops.filter(is_dropoff=True)
    
    # Get upcoming trips passing through this point
    upcoming_trips = Trip.objects.filter(
        route__stops__boarding_point=boarding_point,
        departure_date__gte=timezone.now().date(),
        is_active=True
    ).select_related(
        'bus', 'route', 'route__origin', 'route__destination'
    ).distinct().order_by('departure_date', 'departure_time')[:10]
    
    context = {
        'boarding_point': boarding_point,
        'route_stops': route_stops,
        'pickup_routes': pickup_routes,
        'dropoff_routes': dropoff_routes,
        'upcoming_trips': upcoming_trips,
    }
    
    return render(request, 'admin/boarding_points/boarding_point_detail.html', context)


# ============================================
# ROUTE STOPS VIEWS
# ============================================

def route_stop_list(request):
    """Display all route stops"""
    route_stops = RouteStop.objects.select_related(
        'route', 'route__origin', 'route__destination',
        'boarding_point', 'boarding_point__location'
    ).order_by('route__origin__name', 'stop_order')
    
    # Route filter
    route_id = request.GET.get('route', '')
    if route_id:
        route_stops = route_stops.filter(route_id=route_id)
    
    # Location filter
    location_id = request.GET.get('location', '')
    if location_id:
        route_stops = route_stops.filter(boarding_point__location_id=location_id)
    
    # Stop type filter
    stop_type = request.GET.get('stop_type', '')
    if stop_type == 'pickup':
        route_stops = route_stops.filter(is_pickup=True)
    elif stop_type == 'dropoff':
        route_stops = route_stops.filter(is_dropoff=True)
    
    # Get all routes and locations for filters
    routes = Route.objects.filter(is_active=True).select_related(
        'origin', 'destination'
    ).order_by('origin__name')
    locations = Location.objects.filter(is_active=True).order_by('name')
    
    context = {
        'route_stops': route_stops,
        'routes': routes,
        'locations': locations,
        'selected_route': route_id,
        'selected_location': location_id,
        'selected_stop_type': stop_type,
    }
    
    return render(request, 'admin/route_stops/route_stop_list.html', context)


def route_stop_detail(request, pk):
    """Display detailed route stop information"""
    route_stop = get_object_or_404(
        RouteStop.objects.select_related(
            'route', 'route__origin', 'route__destination',
            'boarding_point', 'boarding_point__location'
        ),
        pk=pk
    )
    
    # Get other stops on the same route
    other_stops = RouteStop.objects.filter(
        route=route_stop.route
    ).exclude(pk=pk).select_related(
        'boarding_point', 'boarding_point__location'
    ).order_by('stop_order')
    
    # Previous and next stops
    previous_stop = RouteStop.objects.filter(
        route=route_stop.route,
        stop_order__lt=route_stop.stop_order
    ).order_by('-stop_order').first()
    
    next_stop = RouteStop.objects.filter(
        route=route_stop.route,
        stop_order__gt=route_stop.stop_order
    ).order_by('stop_order').first()
    
    context = {
        'route_stop': route_stop,
        'other_stops': other_stops,
        'previous_stop': previous_stop,
        'next_stop': next_stop,
    }
    
    return render(request, 'admin/route_stops/route_stop_detail.html', context)



"""
booking/views.py - Complete Admin Views for Navigation Sections
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Count, Sum, Q, Avg, F
from django.db.models.functions import TruncMonth, TruncDate
from django.utils import timezone
from datetime import timedelta, datetime
from django.http import JsonResponse, HttpResponse
import json
from .models import (
    BusOperator, Bus, Location, BoardingPoint, Route, RouteStop,
    Trip, Booking, Payment, Review, Amenity, SeatLayout, Seat, SeatBooking
)


# ==================== CUSTOMERS ====================
def customer_list(request):
    """List all customers with their booking history"""
    # Get unique customers from bookings
    search_query = request.GET.get('search', '')
    
    bookings = Booking.objects.select_related(
        'trip', 'trip__route', 'trip__bus'
    ).order_by('customer_full_name', '-created_at')
    
    if search_query:
        bookings = bookings.filter(
            Q(customer_full_name__icontains=search_query) |
            Q(customer_email__icontains=search_query) |
            Q(customer_phone__icontains=search_query) |
            Q(customer_id_number__icontains=search_query)
        )
    
    # Group customers by email (unique identifier)
    customers_data = {}
    for booking in bookings:
        email = booking.customer_email
        if email not in customers_data:
            customers_data[email] = {
                'full_name': booking.customer_full_name,
                'email': booking.customer_email,
                'phone': booking.customer_phone,
                'id_number': booking.customer_id_number,
                'total_bookings': 0,
                'total_spent': 0,
                'last_booking': None,
                'status': 'active'
            }
        
        customers_data[email]['total_bookings'] += 1
        if booking.status in ['paid', 'confirmed', 'completed']:
            customers_data[email]['total_spent'] += float(booking.total_amount)
        
        if not customers_data[email]['last_booking'] or booking.created_at > customers_data[email]['last_booking']:
            customers_data[email]['last_booking'] = booking.created_at
    
    # Convert to list and sort
    customers = list(customers_data.values())
    customers.sort(key=lambda x: x['total_bookings'], reverse=True)
    
    context = {
        'customers': customers,
        'search_query': search_query,
        'total_customers': len(customers),
    }
    return render(request, 'booking/customers/list.html', context)


def customer_detail(request, email):
    """View detailed customer information and booking history"""
    bookings = Booking.objects.filter(
        customer_email=email
    ).select_related(
        'trip', 'trip__route', 'trip__bus', 'trip__bus__operator'
    ).prefetch_related(
        'seat_bookings__seat', 'payments'
    ).order_by('-created_at')
    
    if not bookings.exists():
        messages.error(request, 'Customer not found')
        return redirect('booking:customer_list')
    
    customer_info = bookings.first()
    
    # Statistics
    total_spent = bookings.filter(
        status__in=['paid', 'confirmed', 'completed']
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    completed_trips = bookings.filter(status='completed').count()
    cancelled_bookings = bookings.filter(status='cancelled').count()
    
    # Favorite route
    favorite_route = bookings.values(
        'trip__route__origin__name',
        'trip__route__destination__name'
    ).annotate(
        count=Count('id')
    ).order_by('-count').first()
    
    context = {
        'customer': {
            'full_name': customer_info.customer_full_name,
            'email': customer_info.customer_email,
            'phone': customer_info.customer_phone,
            'id_number': customer_info.customer_id_number,
        },
        'bookings': bookings,
        'total_bookings': bookings.count(),
        'total_spent': total_spent,
        'completed_trips': completed_trips,
        'cancelled_bookings': cancelled_bookings,
        'favorite_route': favorite_route,
    }
    return render(request, 'booking/customers/detail.html', context)


# ==================== REVIEWS ====================
def review_list(request):
    """List all customer reviews"""
    status_filter = request.GET.get('status', 'all')
    rating_filter = request.GET.get('rating', '')
    search_query = request.GET.get('search', '')
    
    reviews = Review.objects.select_related(
        'booking', 'bus', 'bus__operator', 'booking__trip__route'
    ).order_by('-created_at')
    
    if rating_filter:
        reviews = reviews.filter(rating=rating_filter)
    
    if search_query:
        reviews = reviews.filter(
            Q(booking__customer_full_name__icontains=search_query) |
            Q(bus__bus_name__icontains=search_query) |
            Q(comment__icontains=search_query)
        )
    
    # Statistics
    total_reviews = reviews.count()
    average_rating = reviews.aggregate(avg=Avg('rating'))['avg'] or 0
    
    rating_distribution = {
        5: reviews.filter(rating=5).count(),
        4: reviews.filter(rating=4).count(),
        3: reviews.filter(rating=3).count(),
        2: reviews.filter(rating=2).count(),
        1: reviews.filter(rating=1).count(),
    }
    
    context = {
        'reviews': reviews,
        'total_reviews': total_reviews,
        'average_rating': round(average_rating, 2),
        'rating_distribution': rating_distribution,
        'status_filter': status_filter,
        'rating_filter': rating_filter,
        'search_query': search_query,
    }
    return render(request, 'booking/reviews/list.html', context)


def review_detail(request, pk):
    """View detailed review information"""
    review = get_object_or_404(
        Review.objects.select_related(
            'booking', 'bus', 'bus__operator', 
            'booking__trip', 'booking__trip__route'
        ),
        pk=pk
    )
    
    context = {
        'review': review,
    }
    return render(request, 'booking/reviews/detail.html', context)


"""
booking/views.py - Dynamic Revenue Reports with Export & Analytics
"""
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Sum, Q, Avg, F
from django.db.models.functions import TruncDate, TruncMonth, TruncWeek, TruncDay
from django.utils import timezone
from datetime import timedelta, datetime
import csv
import json
from io import BytesIO
from .models import Payment, Booking, BusOperator, Route, Bus
from django.db.models import Sum, Count
from django.db.models import IntegerField
from django.db.models.expressions import RawSQL

# Optional: For PDF export
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


def revenue_report(request):
    """Dynamic Revenue Analysis with Advanced Filtering and Exports"""
    
    # ==================== FILTERS ====================
    period = request.GET.get('period', 'month')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Additional Filters
    operator_id = request.GET.get('operator')
    route_id = request.GET.get('route')
    payment_method = request.GET.get('payment_method')
    bus_type = request.GET.get('bus_type')
    
    # Export format
    export_format = request.GET.get('export')  # csv, pdf, json
    
    today = timezone.now().date()
    
    # ==================== DATE RANGE ====================
    if period == 'today':
        date_from = today
        date_to = today
    elif period == 'yesterday':
        date_from = today - timedelta(days=1)
        date_to = today - timedelta(days=1)
    elif period == 'week':
        date_from = today - timedelta(days=7)
        date_to = today
    elif period == 'last_week':
        date_from = today - timedelta(days=14)
        date_to = today - timedelta(days=7)
    elif period == 'month':
        date_from = today.replace(day=1)
        date_to = today
    elif period == 'last_month':
        # Get first day of last month
        first_day_this_month = today.replace(day=1)
        last_day_last_month = first_day_this_month - timedelta(days=1)
        date_from = last_day_last_month.replace(day=1)
        date_to = last_day_last_month
    elif period == 'quarter':
        # Current quarter
        current_quarter = (today.month - 1) // 3
        date_from = today.replace(month=current_quarter * 3 + 1, day=1)
        date_to = today
    elif period == 'year':
        date_from = today.replace(month=1, day=1)
        date_to = today
    elif period == 'last_year':
        date_from = today.replace(year=today.year - 1, month=1, day=1)
        date_to = today.replace(year=today.year - 1, month=12, day=31)
    elif period == 'custom' and start_date and end_date:
        try:
            date_from = datetime.strptime(start_date, '%Y-%m-%d').date()
            date_to = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            date_from = today.replace(day=1)
            date_to = today
    else:
        date_from = today.replace(day=1)
        date_to = today
    
    # ==================== BASE QUERY ====================
    payments = Payment.objects.filter(
        status='completed',
        created_at__date__gte=date_from,
        created_at__date__lte=date_to
    ).select_related(
        'booking',
        'booking__trip',
        'booking__trip__bus',
        'booking__trip__bus__operator',
        'booking__trip__route'
    )
    
    # Apply additional filters
    if operator_id:
        payments = payments.filter(booking__trip__bus__operator_id=operator_id)
    
    if route_id:
        payments = payments.filter(booking__trip__route_id=route_id)
    
    if payment_method:
        payments = payments.filter(payment_method=payment_method)
    
    if bus_type:
        payments = payments.filter(booking__trip__bus__bus_type=bus_type)
    
    # ==================== KEY METRICS ====================
    total_revenue = payments.aggregate(total=Sum('amount'))['total'] or 0
    total_transactions = payments.count()
    average_transaction = total_revenue / total_transactions if total_transactions > 0 else 0
    
    # Comparison with previous period
    days_diff = (date_to - date_from).days + 1
    prev_date_from = date_from - timedelta(days=days_diff)
    prev_date_to = date_from - timedelta(days=1)
    
    prev_revenue = Payment.objects.filter(
        status='completed',
        created_at__date__gte=prev_date_from,
        created_at__date__lte=prev_date_to
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    revenue_change = ((total_revenue - prev_revenue) / prev_revenue * 100) if prev_revenue > 0 else 0
    
    # ==================== REVENUE BY PAYMENT METHOD ====================
    revenue_by_method = payments.values('payment_method').annotate(
        total=Sum('amount'),
        count=Count('id'),
        avg=Avg('amount')
    ).order_by('-total')
    
    # Calculate percentages
    for method in revenue_by_method:
        method['percentage'] = (method['total'] / total_revenue * 100) if total_revenue > 0 else 0
    
    # ==================== REVENUE BY OPERATOR ====================
    revenue_by_operator = payments.values(
        operator_name=F('booking__trip__bus__operator__name'),
        operator_id=F('booking__trip__bus__operator_id')
    ).annotate(
        total=Sum('amount'),
        count=Count('id'),
        avg=Avg('amount'),
        buses_used=Count('booking__trip__bus', distinct=True)
    ).order_by('-total')[:10]
    
    for operator in revenue_by_operator:
        operator['percentage'] = (operator['total'] / total_revenue * 100) if total_revenue > 0 else 0
    
    # ==================== REVENUE BY ROUTE ====================
    revenue_by_route = payments.values(
        origin=F('booking__trip__route__origin__name'),
        destination=F('booking__trip__route__destination__name'),
        route_id=F('booking__trip__route_id')
    ).annotate(
        total=Sum('amount'),
        count=Count('id'),
        avg=Avg('amount'),
        distance=F('booking__trip__route__distance_km')
    ).order_by('-total')[:15]
    
    for route in revenue_by_route:
        route['percentage'] = (route['total'] / total_revenue * 100) if total_revenue > 0 else 0
        route['revenue_per_km'] = route['total'] / route['distance'] if route['distance'] else 0
    
    # ==================== REVENUE BY BUS TYPE ====================
    revenue_by_bus_type = payments.values(
        bus_type=F('booking__trip__bus__bus_type')
    ).annotate(
        total=Sum('amount'),
        count=Count('id'),
        avg=Avg('amount')
    ).order_by('-total')
    
    for bus_type_data in revenue_by_bus_type:
        bus_type_data['percentage'] = (bus_type_data['total'] / total_revenue * 100) if total_revenue > 0 else 0
    
    # ==================== TIME-BASED TRENDS ====================
    # Daily revenue (for line chart)
    daily_revenue = payments.annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('date')
    
    # Hourly revenue analysis
    hourly_revenue = payments.annotate(
        hour=RawSQL("CAST(strftime('%%H', created_at) AS INTEGER)", [])
    ).values('hour').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('hour')
    
    # Day of week analysis
    day_of_week_revenue = payments.annotate(
        day=RawSQL("CAST(strftime('%%w', created_at) AS INTEGER)", [])
    ).values('day').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('day')
    
    # Map day numbers to names
    day_names = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    for day_data in day_of_week_revenue:
        day_data['day_name'] = day_names[int(day_data['day'])]

    
    # Monthly trend (last 12 months)
    twelve_months_ago = today - timedelta(days=365)
    monthly_revenue = Payment.objects.filter(
        status='completed',
        created_at__date__gte=twelve_months_ago
    ).annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('month')
    
    # ==================== TOP PERFORMERS ====================
    # Top 10 buses by revenue
    top_buses = payments.values(
        bus_name=F('booking__trip__bus__bus_name'),
        bus_id=F('booking__trip__bus_id'),
        operator_name=F('booking__trip__bus__operator__name')
    ).annotate(
        total=Sum('amount'),
        trips=Count('booking__trip', distinct=True),
        bookings=Count('booking', distinct=True)
    ).order_by('-total')[:10]
    
    # Top customers by spending
    top_customers = payments.values(
        customer_name=F('booking__customer_full_name'),
        customer_email=F('booking__customer_email')
    ).annotate(
        total=Sum('amount'),
        bookings=Count('booking', distinct=True)
    ).order_by('-total')[:10]
    
    # ==================== SEAT CLASS ANALYSIS ====================
    revenue_by_seat_class = payments.values(
        seat_class=F('booking__seat_bookings__seat__seat_class')
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    for seat_class in revenue_by_seat_class:
        seat_class['percentage'] = (seat_class['total'] / total_revenue * 100) if total_revenue > 0 else 0
    
    # ==================== EXPORT HANDLING ====================
    if export_format:
        if export_format == 'csv':
            return export_revenue_csv(
                payments, 
                date_from, 
                date_to, 
                total_revenue, 
                total_transactions,
                revenue_by_method,
                revenue_by_operator,
                revenue_by_route
            )
        elif export_format == 'json':
            return export_revenue_json(
                date_from,
                date_to,
                total_revenue,
                total_transactions,
                revenue_by_method,
                revenue_by_operator,
                revenue_by_route,
                daily_revenue
            )
        elif export_format == 'pdf' and REPORTLAB_AVAILABLE:
            return export_revenue_pdf(
                date_from,
                date_to,
                total_revenue,
                total_transactions,
                revenue_by_method,
                revenue_by_operator,
                revenue_by_route
            )
    
    # ==================== PREPARE CHART DATA ====================
    # Format data for Chart.js
    chart_data = {
        'daily_labels': [item['date'].strftime('%Y-%m-%d') for item in daily_revenue],
        'daily_revenue': [float(item['total']) for item in daily_revenue],
        'daily_transactions': [item['count'] for item in daily_revenue],
        
        'method_labels': [method['payment_method'].upper() for method in revenue_by_method],
        'method_data': [float(method['total']) for method in revenue_by_method],
        
        'operator_labels': [op['operator_name'] for op in revenue_by_operator],
        'operator_data': [float(op['total']) for op in revenue_by_operator],
        
        'hourly_labels': [f"{int(item['hour']):02d}:00" for item in hourly_revenue],
        'hourly_data': [float(item['total']) for item in hourly_revenue],
        
        'day_labels': [item['day_name'] for item in day_of_week_revenue],
        'day_data': [float(item['total']) for item in day_of_week_revenue],
        
        'seat_class_labels': [sc['seat_class'] or 'Unknown' for sc in revenue_by_seat_class],
        'seat_class_data': [float(sc['total']) for sc in revenue_by_seat_class],
    }
    
    # ==================== FILTER OPTIONS ====================
    operators = BusOperator.objects.filter(is_active=True).order_by('name')
    routes = Route.objects.filter(is_active=True).select_related('origin', 'destination').order_by('origin__name')
    payment_methods = [choice[0] for choice in Payment.PAYMENT_METHOD_CHOICES]
    bus_types = [choice[0] for choice in Bus.BUS_TYPE_CHOICES]
    
    context = {
        # Filters
        'period': period,
        'date_from': date_from,
        'date_to': date_to,
        'selected_operator': operator_id,
        'selected_route': route_id,
        'selected_payment_method': payment_method,
        'selected_bus_type': bus_type,
        
        # Filter options
        'operators': operators,
        'routes': routes,
        'payment_methods': payment_methods,
        'bus_types': bus_types,
        
        # Key metrics
        'total_revenue': total_revenue,
        'total_transactions': total_transactions,
        'average_transaction': average_transaction,
        'prev_revenue': prev_revenue,
        'revenue_change': revenue_change,
        
        # Analysis data
        'revenue_by_method': revenue_by_method,
        'revenue_by_operator': revenue_by_operator,
        'revenue_by_route': revenue_by_route,
        'revenue_by_bus_type': revenue_by_bus_type,
        'revenue_by_seat_class': revenue_by_seat_class,
        
        # Trends
        'daily_revenue': daily_revenue,
        'hourly_revenue': hourly_revenue,
        'day_of_week_revenue': day_of_week_revenue,
        'monthly_revenue': monthly_revenue,
        
        # Top performers
        'top_buses': top_buses,
        'top_customers': top_customers,
        
        # Chart data (JSON)
        'chart_data': json.dumps(chart_data),
    }
    
    return render(request, 'booking/reports/revenue.html', context)


# ==================== EXPORT FUNCTIONS ====================

def export_revenue_csv(payments, date_from, date_to, total_revenue, total_transactions, 
                       revenue_by_method, revenue_by_operator, revenue_by_route):
    """Export revenue report as CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="revenue_report_{date_from}_{date_to}.csv"'
    
    writer = csv.writer(response)
    
    # Header
    writer.writerow(['Revenue Report'])
    writer.writerow(['Period', f'{date_from} to {date_to}'])
    writer.writerow(['Generated', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow([])
    
    # Summary
    writer.writerow(['SUMMARY'])
    writer.writerow(['Total Revenue', f'KES {total_revenue:,.2f}'])
    writer.writerow(['Total Transactions', total_transactions])
    writer.writerow(['Average Transaction', f'KES {total_revenue/total_transactions:,.2f}' if total_transactions > 0 else 'N/A'])
    writer.writerow([])
    
    # Revenue by Payment Method
    writer.writerow(['REVENUE BY PAYMENT METHOD'])
    writer.writerow(['Payment Method', 'Transactions', 'Total Revenue', 'Percentage'])
    for method in revenue_by_method:
        writer.writerow([
            method['payment_method'].upper(),
            method['count'],
            f"KES {method['total']:,.2f}",
            f"{method['percentage']:.2f}%"
        ])
    writer.writerow([])
    
    # Revenue by Operator
    writer.writerow(['REVENUE BY OPERATOR'])
    writer.writerow(['Operator', 'Bookings', 'Total Revenue', 'Average', 'Percentage'])
    for operator in revenue_by_operator:
        writer.writerow([
            operator['operator_name'],
            operator['count'],
            f"KES {operator['total']:,.2f}",
            f"KES {operator['avg']:,.2f}",
            f"{operator['percentage']:.2f}%"
        ])
    writer.writerow([])
    
    # Revenue by Route
    writer.writerow(['REVENUE BY ROUTE'])
    writer.writerow(['Origin', 'Destination', 'Bookings', 'Total Revenue', 'Average', 'Percentage'])
    for route in revenue_by_route:
        writer.writerow([
            route['origin'],
            route['destination'],
            route['count'],
            f"KES {route['total']:,.2f}",
            f"KES {route['avg']:,.2f}",
            f"{route['percentage']:.2f}%"
        ])
    
    return response


def export_revenue_json(date_from, date_to, total_revenue, total_transactions,
                        revenue_by_method, revenue_by_operator, revenue_by_route, daily_revenue):
    """Export revenue report as JSON"""
    data = {
        'report_info': {
            'period_start': str(date_from),
            'period_end': str(date_to),
            'generated_at': timezone.now().isoformat(),
        },
        'summary': {
            'total_revenue': float(total_revenue),
            'total_transactions': total_transactions,
            'average_transaction': float(total_revenue / total_transactions) if total_transactions > 0 else 0,
        },
        'revenue_by_method': [
            {
                'payment_method': m['payment_method'],
                'transactions': m['count'],
                'total': float(m['total']),
                'percentage': float(m['percentage'])
            } for m in revenue_by_method
        ],
        'revenue_by_operator': [
            {
                'operator': o['operator_name'],
                'bookings': o['count'],
                'total': float(o['total']),
                'average': float(o['avg']),
                'percentage': float(o['percentage'])
            } for o in revenue_by_operator
        ],
        'revenue_by_route': [
            {
                'origin': r['origin'],
                'destination': r['destination'],
                'bookings': r['count'],
                'total': float(r['total']),
                'average': float(r['avg']),
                'percentage': float(r['percentage'])
            } for r in revenue_by_route
        ],
        'daily_trend': [
            {
                'date': str(d['date']),
                'revenue': float(d['total']),
                'transactions': d['count']
            } for d in daily_revenue
        ]
    }
    
    response = HttpResponse(json.dumps(data, indent=2), content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename="revenue_report_{date_from}_{date_to}.json"'
    return response


def export_revenue_pdf(date_from, date_to, total_revenue, total_transactions,
                       revenue_by_method, revenue_by_operator, revenue_by_route):
    """Export revenue report as PDF (requires reportlab)"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"Revenue Report: {date_from} to {date_to}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Summary table
    summary_data = [
        ['Metric', 'Value'],
        ['Total Revenue', f'KES {total_revenue:,.2f}'],
        ['Total Transactions', f'{total_transactions}'],
        ['Average Transaction', f'KES {total_revenue/total_transactions:,.2f}' if total_transactions > 0 else 'N/A'],
    ]
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="revenue_report_{date_from}_{date_to}.pdf"'
    return response

def booking_report(request):
    """Booking statistics and analysis"""
    period = request.GET.get('period', 'month')
    today = timezone.now().date()
    
    # Determine date range
    if period == 'day':
        date_from = today
    elif period == 'week':
        date_from = today - timedelta(days=7)
    elif period == 'month':
        date_from = today - timedelta(days=30)
    else:
        date_from = today - timedelta(days=90)
    
    bookings = Booking.objects.filter(created_at__date__gte=date_from)
    
    # Statistics
    total_bookings = bookings.count()
    confirmed_bookings = bookings.filter(status__in=['confirmed', 'paid', 'completed']).count()
    pending_bookings = bookings.filter(status='pending').count()
    cancelled_bookings = bookings.filter(status='cancelled').count()
    
    # Booking status distribution
    status_distribution = bookings.values('status').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Bookings by route
    bookings_by_route = bookings.values(
        'trip__route__origin__name',
        'trip__route__destination__name'
    ).annotate(
        count=Count('id'),
        revenue=Sum('total_amount')
    ).order_by('-count')[:10]
    
    # Bookings by operator
    bookings_by_operator = bookings.values(
        'trip__bus__operator__name'
    ).annotate(
        count=Count('id'),
        revenue=Sum('total_amount')
    ).order_by('-count')
    
    # Daily booking trend
    daily_bookings = bookings.annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        count=Count('id')
    ).order_by('date')
    
    # Peak booking times
    hourly_bookings = bookings.extra(
        select={'hour': "EXTRACT(hour FROM created_at)"}
    ).values('hour').annotate(
        count=Count('id')
    ).order_by('hour')
    
    # Seat class preferences
    seat_class_stats = SeatBooking.objects.filter(
        booking__in=bookings
    ).values('seat__seat_class').annotate(
        count=Count('id'),
        revenue=Sum('fare')
    ).order_by('-count')
    
    # Average booking value
    avg_booking_value = bookings.aggregate(avg=Avg('total_amount'))['avg'] or 0
    
    context = {
        'period': period,
        'date_from': date_from,
        'total_bookings': total_bookings,
        'confirmed_bookings': confirmed_bookings,
        'pending_bookings': pending_bookings,
        'cancelled_bookings': cancelled_bookings,
        'status_distribution': status_distribution,
        'bookings_by_route': bookings_by_route,
        'bookings_by_operator': bookings_by_operator,
        'daily_bookings': list(daily_bookings),
        'hourly_bookings': list(hourly_bookings),
        'seat_class_stats': seat_class_stats,
        'avg_booking_value': round(avg_booking_value, 2),
    }
    return render(request, 'booking/reports/bookings.html', context)


def analytics_dashboard(request):
    """Comprehensive analytics dashboard"""
    today = timezone.now().date()
    
    # Time period filters
    last_30_days = today - timedelta(days=30)
    last_7_days = today - timedelta(days=7)
    
    # Key Performance Indicators
    total_revenue = Payment.objects.filter(
        status='completed'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    revenue_last_30 = Payment.objects.filter(
        status='completed',
        created_at__date__gte=last_30_days
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    total_bookings = Booking.objects.count()
    bookings_last_30 = Booking.objects.filter(
        created_at__date__gte=last_30_days
    ).count()
    
    total_customers = Booking.objects.values('customer_email').distinct().count()
    
    # Occupancy rate
    total_trips = Trip.objects.filter(
        departure_date__gte=last_30_days,
        status__in=['completed', 'departed']
    ).count()
    
    if total_trips > 0:
        trips_data = Trip.objects.filter(
            departure_date__gte=last_30_days,
            status__in=['completed', 'departed']
        ).annotate(
            total_seats=F('bus__seat_layout__total_seats'),
            booked_seats=Count('bookings__seat_bookings')
        )
        
        total_capacity = sum(t.total_seats for t in trips_data)
        total_booked = sum(t.booked_seats for t in trips_data)
        occupancy_rate = (total_booked / total_capacity * 100) if total_capacity > 0 else 0
    else:
        occupancy_rate = 0
    
    # Top performing routes
    top_routes = Route.objects.annotate(
        booking_count=Count('trips__bookings', filter=Q(
            trips__bookings__created_at__date__gte=last_30_days
        )),
        revenue=Sum('trips__bookings__total_amount', filter=Q(
            trips__bookings__created_at__date__gte=last_30_days,
            trips__bookings__status__in=['paid', 'confirmed', 'completed']
        ))
    ).order_by('-booking_count')[:5]
    
    # Top performing buses
    top_buses = Bus.objects.annotate(
        trip_count=Count('trips', filter=Q(
            trips__departure_date__gte=last_30_days
        )),
        revenue=Sum('trips__bookings__total_amount', filter=Q(
            trips__bookings__created_at__date__gte=last_30_days,
            trips__bookings__status__in=['paid', 'confirmed', 'completed']
        ))
    ).order_by('-revenue')[:5]
    
    # Customer satisfaction (average rating)
    avg_rating = Review.objects.aggregate(avg=Avg('rating'))['avg'] or 0
    recent_reviews = Review.objects.select_related(
        'bus', 'booking'
    ).order_by('-created_at')[:5]
    
    # Revenue trend (last 7 days)
    revenue_trend = Payment.objects.filter(
        status='completed',
        created_at__date__gte=last_7_days
    ).annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        total=Sum('amount')
    ).order_by('date')
    
    # Booking trend (last 7 days)
    booking_trend = Booking.objects.filter(
        created_at__date__gte=last_7_days
    ).annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        count=Count('id')
    ).order_by('date')
    
    # Most popular amenities
    popular_amenities = Amenity.objects.annotate(
        bus_count=Count('bus')
    ).order_by('-bus_count')[:5]
    
    context = {
        'total_revenue': total_revenue,
        'revenue_last_30': revenue_last_30,
        'total_bookings': total_bookings,
        'bookings_last_30': bookings_last_30,
        'total_customers': total_customers,
        'occupancy_rate': round(occupancy_rate, 2),
        'avg_rating': round(avg_rating, 2),
        'top_routes': top_routes,
        'top_buses': top_buses,
        'recent_reviews': recent_reviews,
        'revenue_trend': list(revenue_trend),
        'booking_trend': list(booking_trend),
        'popular_amenities': popular_amenities,
    }
    return render(request, 'booking/reports/analytics.html', context)


# ==================== AMENITIES ====================
def amenity_list(request):
    """List all bus amenities"""
    amenities = Amenity.objects.annotate(
        bus_count=Count('bus')
    ).order_by('name')
    
    context = {
        'amenities': amenities,
    }
    return render(request, 'booking/amenities/list.html', context)


def amenity_create(request):
    """Create new amenity"""
    if request.method == 'POST':
        name = request.POST.get('name')
        icon = request.POST.get('icon')
        description = request.POST.get('description', '')
        
        amenity = Amenity.objects.create(
            name=name,
            icon=icon,
            description=description
        )
        
        messages.success(request, f'Amenity "{name}" created successfully!')
        return redirect('booking:amenity_list')
    
    return render(request, 'booking/amenities/create.html')


def amenity_edit(request, pk):
    """Edit amenity"""
    amenity = get_object_or_404(Amenity, pk=pk)
    
    if request.method == 'POST':
        amenity.name = request.POST.get('name')
        amenity.icon = request.POST.get('icon')
        amenity.description = request.POST.get('description', '')
        amenity.save()
        
        messages.success(request, 'Amenity updated successfully!')
        return redirect('booking:amenity_list')
    
    context = {
        'amenity': amenity,
    }
    return render(request, 'booking/amenities/edit.html', context)


def amenity_delete(request, pk):
    """Delete amenity"""
    amenity = get_object_or_404(Amenity, pk=pk)
    
    if request.method == 'POST':
        name = amenity.name
        amenity.delete()
        messages.success(request, f'Amenity "{name}" deleted successfully!')
        return redirect('booking:amenity_list')
    
    context = {
        'amenity': amenity,
    }
    return render(request, 'booking/amenities/delete.html', context)


# ==================== SETTINGS ====================
def settings_view(request):
    """System settings page"""
    if request.method == 'POST':
        # Handle settings update
        messages.success(request, 'Settings updated successfully!')
        return redirect('booking:settings')
    
    # Get system statistics
    stats = {
        'total_operators': BusOperator.objects.filter(is_active=True).count(),
        'total_buses': Bus.objects.filter(is_active=True).count(),
        'total_routes': Route.objects.filter(is_active=True).count(),
        'total_locations': Location.objects.filter(is_active=True).count(),
        'total_amenities': Amenity.objects.count(),
        'total_seat_layouts': SeatLayout.objects.count(),
    }
    
    context = {
        'stats': stats,
    }
    return render(request, 'booking/settings/index.html', context)